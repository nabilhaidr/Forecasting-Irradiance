"""Parse operator workbooks for maintenance, outage, and curtailment periods.

These workbooks are operational corroboration for the sensor-derived coverage
audit, not authoritative telemetry. Every parser is deterministic, anchors on
header text rather than fixed cell coordinates so it survives the per-era
layout drift in the source files, and carries file-level provenance so each
claim can be traced back to a byte range and hash.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

from .cov_ingest import sha256_file


# Indonesian (and a few English) month tokens used in the operator workbooks'
# text-formatted date cells, e.g. "1-Mei-25", "15-Agu-25", "3-Des-25".
_MONTH_TOKENS: dict[str, int] = {
    "jan": 1, "januari": 1,
    "feb": 2, "peb": 2, "februari": 2, "februari.": 2,
    "mar": 3, "mrt": 3, "maret": 3,
    "apr": 4, "april": 4,
    "mei": 5, "may": 5,
    "jun": 6, "juni": 6,
    "jul": 7, "juli": 7,
    "agu": 8, "ags": 8, "agt": 8, "agust": 8, "agustus": 8, "aug": 8,
    "sep": 9, "sept": 9, "september": 9,
    "okt": 10, "oct": 10, "oktober": 10,
    "nov": 11, "nop": 11, "november": 11,
    "des": 12, "dec": 12, "desember": 12,
}
_TEXT_DATE = re.compile(r"^\s*(\d{1,2})[-/ ]([A-Za-z]+)[-/ ](\d{2,4})\s*$")


def _parse_text_date(text: str) -> date | None:
    match = _TEXT_DATE.match(text)
    if match is None:
        return None
    day = int(match.group(1))
    month = _MONTH_TOKENS.get(match.group(2).strip().lower())
    if month is None:
        return None
    year = int(match.group(3))
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def file_provenance(path: Path) -> dict[str, object]:
    """Return name, byte size, and SHA-256 for a source workbook."""

    path = Path(path)
    return {
        "source_name": path.name,
        "byte_size": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def _sheet_rows(path: Path, sheet: str | None) -> list[tuple[str, list[object]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = workbook.sheetnames if sheet is None else [sheet]
        result: list[tuple[str, list[object]]] = []
        for name in names:
            worksheet = workbook[name]
            worksheet.reset_dimensions()
            for row in worksheet.iter_rows(values_only=True):
                result.append((name, list(row)))
        return result
    finally:
        workbook.close()


def _as_str(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _as_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return _parse_text_date(value)
    return None


def _locate_header(rows: list[list[object]], anchor: str) -> int | None:
    anchor_lower = anchor.lower()
    for index, row in enumerate(rows):
        for cell in row:
            if isinstance(cell, str) and anchor_lower in cell.lower():
                return index
    return None


def _column_index(header: list[object], label: str) -> int | None:
    label_lower = label.lower()
    for index, cell in enumerate(header):
        if isinstance(cell, str) and cell.strip().lower() == label_lower:
            return index
    return None


def parse_maintenance_records(path: Path) -> pd.DataFrame:
    """Extract every maintenance-log row across all monthly sheets."""

    path = Path(path)
    grouped: dict[str, list[list[object]]] = {}
    for sheet_name, row in _sheet_rows(path, None):
        grouped.setdefault(sheet_name, []).append(row)

    records: list[dict[str, object]] = []
    for sheet_name, rows in grouped.items():
        header_index = _locate_header(rows, "Type of Maintenance")
        if header_index is None:
            continue
        header = rows[header_index]
        cols = {
            "type": _column_index(header, "Type of Maintenance"),
            "area": _column_index(header, "Area"),
            "equipment": _column_index(header, "Equipment"),
            "trouble": _column_index(header, "Trouble Description"),
            "action": _column_index(header, "Comment/Action"),
            "start": _column_index(header, "Start"),
            "stop": _column_index(header, "Stop"),
            "status": _column_index(header, "Progress/Status"),
        }
        date_col = 0

        def cell(row: list[object], key: str) -> object:
            index = cols[key]
            if index is None or index >= len(row):
                return None
            return row[index]

        last_date: date | None = None
        for row in rows[header_index + 1 :]:
            row_date = _as_date(row[date_col]) if row else None
            if row_date is not None:
                last_date = row_date
            type_value = _as_str(cell(row, "type"))
            equipment = _as_str(cell(row, "equipment"))
            trouble = _as_str(cell(row, "trouble"))
            if type_value is None and equipment is None and trouble is None:
                continue
            haystack = " ".join(
                filter(None, [equipment, _as_str(cell(row, "area"))])
            ).lower()
            is_ws = "ws-" in haystack or "weather station" in haystack
            records.append(
                {
                    "source_sheet": sheet_name,
                    "date": last_date,
                    "maintenance_type": type_value,
                    "area": _as_str(cell(row, "area")),
                    "equipment": equipment,
                    "trouble_description": trouble,
                    "action": _as_str(cell(row, "action")),
                    "start_time": _as_str(cell(row, "start")),
                    "stop_time": _as_str(cell(row, "stop")),
                    "status": _as_str(cell(row, "status")),
                    "is_weather_station": is_ws,
                }
            )
    return pd.DataFrame(
        records,
        columns=[
            "source_sheet",
            "date",
            "maintenance_type",
            "area",
            "equipment",
            "trouble_description",
            "action",
            "start_time",
            "stop_time",
            "status",
            "is_weather_station",
        ],
    )


def parse_curtailment_daily(path: Path, *, sheet: str) -> pd.DataFrame:
    """Extract positive daily curtailed-energy rows from a generation sheet."""

    path = Path(path)
    rows = [row for _, row in _sheet_rows(path, sheet)]
    header_index = _locate_header(rows, "Curtailed Energy")
    if header_index is None:
        return pd.DataFrame(
            columns=["source_sheet", "date", "curtailed_energy_kwh"]
        )
    header = rows[header_index]
    date_col = _column_index(header, "Date")
    curtailed_col = _column_index(header, "Curtailed Energy (kWh)")
    records: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        if date_col is None or curtailed_col is None:
            break
        row_date = _as_date(row[date_col]) if date_col < len(row) else None
        value = row[curtailed_col] if curtailed_col < len(row) else None
        if row_date is None or not isinstance(value, (int, float)):
            continue
        if value <= 0:
            continue
        records.append(
            {
                "source_sheet": sheet,
                "date": row_date,
                "curtailed_energy_kwh": float(value),
            }
        )
    return pd.DataFrame(
        records, columns=["source_sheet", "date", "curtailed_energy_kwh"]
    )


def _span_bounds(top_header: list[object], label: str) -> int | None:
    label_lower = label.lower()
    for index, cell in enumerate(top_header):
        if isinstance(cell, str) and label_lower in cell.lower():
            return index
    return None


def _find_in_range(
    sub_header: list[object], label: str, start: int, end: int
) -> int | None:
    label_lower = label.lower()
    for index in range(start, min(end, len(sub_header))):
        cell = sub_header[index]
        if isinstance(cell, str) and cell.strip().lower() == label_lower:
            return index
    return None


def parse_dcm_outage_limitation(path: Path, *, sheet: str) -> pd.DataFrame:
    """Extract outage and limitation intervals from a DCM monthly sheet."""

    path = Path(path)
    rows = [row for _, row in _sheet_rows(path, sheet)]
    sub_index = _locate_header(rows, "Restoration Time")
    if sub_index is None or sub_index == 0:
        return pd.DataFrame(
            columns=[
                "source_sheet",
                "date",
                "kind",
                "start_time",
                "restoration_time",
                "equipment",
                "panel",
                "minute_lost",
            ]
        )
    top_header = rows[sub_index - 1]
    sub_header = rows[sub_index]
    outage_start = _span_bounds(top_header, "External Outage")
    limitation_start = _span_bounds(top_header, "External Limitation")
    spans: list[tuple[str, int, int]] = []
    if outage_start is not None:
        end = limitation_start if limitation_start is not None else len(sub_header)
        spans.append(("outage", outage_start, end))
    if limitation_start is not None:
        spans.append(("limitation", limitation_start, len(sub_header)))

    records: list[dict[str, object]] = []
    last_date: date | None = None
    for row in rows[sub_index + 1 :]:
        row_date = _as_date(row[0]) if row else None
        if row_date is not None:
            last_date = row_date
        for kind, start, end in spans:
            start_col = _find_in_range(sub_header, "Start Time", start, end)
            restore_col = _find_in_range(sub_header, "Restoration Time", start, end)
            equip_col = _find_in_range(sub_header, "Equipment Name", start, end)
            panel_col = _find_in_range(sub_header, "Panel Number", start, end)
            minute_col = _find_in_range(sub_header, "Minute Lost", start, end)
            if minute_col is None:
                minute_col = _find_in_range(sub_header, "Minute", start, end)

            def value_at(col: int | None) -> object:
                if col is None or col >= len(row):
                    return None
                return row[col]

            start_time = _as_str(value_at(start_col))
            if start_time is None:
                continue
            minute_value = value_at(minute_col)
            records.append(
                {
                    "source_sheet": sheet,
                    "date": last_date,
                    "kind": kind,
                    "start_time": start_time,
                    "restoration_time": _as_str(value_at(restore_col)),
                    "equipment": _as_str(value_at(equip_col)),
                    "panel": _as_str(value_at(panel_col)),
                    "minute_lost": (
                        int(minute_value)
                        if isinstance(minute_value, (int, float))
                        else None
                    ),
                }
            )
    return pd.DataFrame(
        records,
        columns=[
            "source_sheet",
            "date",
            "kind",
            "start_time",
            "restoration_time",
            "equipment",
            "panel",
            "minute_lost",
        ],
    )
