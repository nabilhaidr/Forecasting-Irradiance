from __future__ import annotations

import json
from datetime import datetime, time
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import yaml

from src.characterisation.data_audit_cli import (
    run_irradiance_audit,
    run_operational_audit,
)


def _site_config(path: Path) -> None:
    config = {
        "site": {
            "latitude_deg": -0.9911713,
            "longitude_deg": 116.6381113,
            "elevation_m": 85.0,
            "timezone": "Asia/Makassar",
            "canonical_freq": "1min",
        }
    }
    path.write_text(yaml.safe_dump(config), encoding="utf-8")


def _ghi_workbook(path: Path, day: str) -> None:
    times = pd.date_range(f"{day} 06:00", f"{day} 18:00", freq="5min")
    zenith = np.abs(np.linspace(-1, 1, len(times)))
    values = np.round(900.0 * (1 - zenith**2), 1)
    frame = pd.DataFrame(
        {
            "row": range(len(times)),
            "date_time": times,
            "GHI_PLTS-IKN_WS-1": values,
            "object_caeid": ["0"] * len(times),
        }
    )
    frame.to_excel(path, index=False)


def _operational_workbooks(spec_dir: Path) -> None:
    maint = openpyxl.Workbook()
    ws = maint.active
    ws.title = "Maret 2025"
    ws.append([None] * 10)
    ws.append([None] * 10)
    ws.append([None] * 10)
    ws.append([None] * 10)
    ws.append([None] * 10)
    ws.append(
        [
            "Date",
            "Type of Maintenance",
            "Area",
            "Equipment",
            "Trouble Description",
            "Comment/Action",
            "Start",
            "Stop",
            "PIC",
            "Progress/Status",
        ]
    )
    ws.append(
        [
            datetime(2025, 3, 6),
            "Corrective",
            "STS 6",
            "Weather Station 3 (WS-3)",
            "Communication Loss",
            "Check",
            time(6, 0),
            time(18, 0),
            "Veri",
            "Open",
        ]
    )
    maint.save(spec_dir / "Maintenance Record PLTS IKN 50 MW.xlsx")

    gen = openpyxl.Workbook()
    gs = gen.active
    gs.title = "Detail2"
    gs.append(["Details for Sum of Total STS"])
    gs.append([None])
    gs.append(
        [
            "Date",
            "Energy (kWh)",
            "GHI",
            "Total Meter Generation (kWh)",
            "Curtailed Energy (kWh)",
        ]
    )
    gs.append([datetime(2025, 1, 31), 1.0, 4.2, 164705.6, 47406.89])
    gen.save(spec_dir / "IKN Generation.xlsx")

    dcm = openpyxl.Workbook()
    ds = dcm.active
    ds.title = "Juli 2025"
    ds.append(["Date", "External Outage", None, None, None, "External Limitation"])
    ds.append(
        [
            None,
            "Start Time",
            "Restoration Time",
            "Equipment Name",
            "Panel Number",
            "Start Time",
            "Restoration Time",
            "Equipment Name",
            "Panel Number",
        ]
    )
    ds.append(
        [
            datetime(2025, 7, 1),
            "06:00:00",
            "15:04:00",
            "OG.3",
            "J14",
            "06:00:00",
            "17:03:00",
            "STS 6 - STS 10",
            "J10",
        ]
    )
    dcm.save(spec_dir / "DCM Manual Calucation Rekon PLTS IKN.xlsx")


def test_run_operational_audit_writes_artifacts(tmp_path: Path) -> None:
    spec_dir = tmp_path / "spec_raw"
    spec_dir.mkdir()
    _operational_workbooks(spec_dir)
    output_dir = tmp_path / "out"

    summary = run_operational_audit(spec_raw_dir=spec_dir, output_dir=output_dir)

    for name in (
        "maintenance_periods.csv",
        "curtailment_periods.csv",
        "dcm_outage_limitation.csv",
        "operational_source_manifest.csv",
        "run_manifest.json",
    ):
        assert (output_dir / name).is_file(), name
    manifest = json.loads((output_dir / "run_manifest.json").read_text())
    assert "maintenance_periods.csv" in manifest["artifact_sha256"]
    assert all(len(v) == 64 for v in manifest["artifact_sha256"].values())
    assert summary["maintenance_row_count"] == 1
    assert summary["curtailment_row_count"] == 1
    maintenance = pd.read_csv(output_dir / "maintenance_periods.csv")
    assert bool(maintenance.iloc[0]["is_weather_station"])


def test_run_irradiance_audit_writes_coverage_and_kc(tmp_path: Path) -> None:
    xlsx_root = tmp_path / "xlsx"
    xlsx_root.mkdir()
    _ghi_workbook(xlsx_root / "GHI_PLTS-IKN_WS-1_2026-Jan.xlsx", "2026-01-05")
    _ghi_workbook(xlsx_root / "GHI_PLTS-IKN_WS-1_2026-Jan-06.xlsx", "2026-01-06")
    config_path = tmp_path / "site.yaml"
    _site_config(config_path)
    output_dir = tmp_path / "out"

    summary = run_irradiance_audit(
        xlsx_root=xlsx_root,
        cov_dir=None,
        site_config_path=config_path,
        output_dir=output_dir,
    )

    for name in (
        "source_manifest.csv",
        "coverage_daily.csv",
        "coverage_monthly.csv",
        "gap_profile.csv",
        "kc_monthly.csv",
        "regime_monthly.csv",
        "outage_candidates.csv",
        "operator_lead_crosscheck.csv",
        "run_manifest.json",
    ):
        assert (output_dir / name).is_file(), name
    manifest = json.loads((output_dir / "run_manifest.json").read_text())
    assert manifest["timezone_caveat"].startswith("historian")
    monthly = pd.read_csv(output_dir / "coverage_monthly.csv")
    assert (monthly["channel_key"] == "GHI").any()
    kc = pd.read_csv(output_dir / "kc_monthly.csv")
    assert (kc["kc_p50"] > 0).all()
    assert summary["coverage_start"].startswith("2026-01-05")
