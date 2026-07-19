from __future__ import annotations

from datetime import datetime, time
from pathlib import Path

import openpyxl

from src.characterisation.operational_periods import (
    file_provenance,
    parse_curtailment_daily,
    parse_dcm_outage_limitation,
    parse_maintenance_records,
)


def _maintenance_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maret 2025"
    ws["E1"] = "Maintenance Record PLTS IKN"
    ws["A3"] = "PLTS"
    ws["A4"] = "Contractor"
    header = [
        "Date",
        "Type of Maintenance",
        "Area",
        "Equipment",
        "Trouble Description",
        "Comment/Action",
        "Start",
        "Stop",
        "PIC",
        "Progress/Status",
        "Date",
        None,
    ]
    ws.append([None] * 12)  # row5 spacer so header lands on row6
    ws.append(header)
    ws.append([None] * 9 + ["Open", "Close"])  # sub-header row7
    ws.append(
        [
            datetime(2025, 3, 6),
            "Corrective",
            "STS 6 & Powerhouse",
            "Weather Station 3 (WS-3)",
            "Communication Loss",
            "Reading condition on SCADA",
            "06:00:00",
            "18:00:00",
            "Veri",
            "Open",
            datetime(2025, 3, 6),
            datetime(2025, 3, 31),
        ]
    )
    ws.append(
        [
            None,  # blank date -> ffill
            "Preventive",
            "PV Module",
            "PV Table",
            "Mowing and Herbicide",
            "Cut plants",
            "09:00:00",
            "16:00:00",
            "Alex",
            "Close",
            datetime(2025, 3, 6),
            datetime(2025, 3, 6),
        ]
    )
    wb.save(path)


def _generation_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail2"
    ws["A1"] = "Details for Sum of Total STS"
    ws.append([None])  # row2 spacer
    ws.append(
        [
            "Date",
            "Energy (kWh)",
            "GHI",
            "SY",
            "PR",
            "Availability",
            "Total Meter Generation (kWh)",
            "Curtailed Energy (kWh)",
        ]
    )
    ws.append([datetime(2025, 1, 31), 1.0, 4.2, 3.4, 0.8, 0.99, 164705.6, 47406.89])
    ws.append([datetime(2025, 1, 30), 1.0, 4.2, 3.4, 0.8, 0.99, 144636.6, 0.0])
    wb.save(path)


def _dcm_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Juli 2025"
    top = ["Date", "External Outage", None, None, None, None, None, None,
           "External Limitation"]
    sub = [
        None,
        "Start Time",
        "Start Time (Rekon)",
        "Restoration Time (Rekon)",
        "Restoration Time",
        "Equipment Name",
        "Panel Number",
        "Minute Lost",
        "Start Time",
        "Restoration Time",
        "Equipment Name",
        "Panel Number",
        "Minute",
    ]
    ws.append(top)
    ws.append(sub)
    ws.append(
        [
            datetime(2025, 7, 1),
            "06:00:00",
            "06:00:00",
            "15:04:00",
            "15:04:00",
            "OG.3",
            "J14",
            544,
            "06:00:00",
            "17:03:00",
            "STS 6 - STS 10",
            "J10, J11, J12",
            663,
        ]
    )
    ws.append([datetime(2025, 7, 2)] + [None] * 12)  # empty day
    wb.save(path)


def test_parse_maintenance_records_ffill_and_ws_flag(tmp_path: Path) -> None:
    path = tmp_path / "maintenance.xlsx"
    _maintenance_workbook(path)

    records = parse_maintenance_records(path)

    assert list(records["date"]) == [
        datetime(2025, 3, 6).date(),
        datetime(2025, 3, 6).date(),
    ]
    ws3 = records.iloc[0]
    assert ws3["maintenance_type"] == "Corrective"
    assert ws3["equipment"] == "Weather Station 3 (WS-3)"
    assert ws3["trouble_description"] == "Communication Loss"
    assert ws3["start_time"] == "06:00:00"
    assert ws3["status"] == "Open"
    assert bool(ws3["is_weather_station"]) is True
    assert bool(records.iloc[1]["is_weather_station"]) is False
    assert (records["source_sheet"] == "Maret 2025").all()


def _maintenance_text_date_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mei 2025"
    header = [
        "Date",
        "Type of Maintenance",
        "Area",
        "Equipment",
        "Trouble Description",
        "Comment/Action",
        "Start",
        "Stop",
        "PIC",
        "Progress/Status",
    ]
    ws.append([None] * 10)  # row5 spacer so header lands on row6
    ws.append(header)
    ws.append(
        [
            "1-Mei-25",  # Indonesian text date
            "Corrective",
            "STS 6",
            "Weather Station 3 (WS-3)",
            "Communication Loss",
            "Check",
            time(6, 30),  # time object, not string
            time(16, 0),
            "Veri",
            "Open",
        ]
    )
    wb.save(path)


def test_parse_maintenance_records_indonesian_text_dates(tmp_path: Path) -> None:
    path = tmp_path / "maintenance_text.xlsx"
    _maintenance_text_date_workbook(path)

    records = parse_maintenance_records(path)

    assert records.iloc[0]["date"] == datetime(2025, 5, 1).date()
    assert records.iloc[0]["start_time"] == "06:30:00"
    assert bool(records.iloc[0]["is_weather_station"]) is True


def test_parse_curtailment_daily_positive_only(tmp_path: Path) -> None:
    path = tmp_path / "generation.xlsx"
    _generation_workbook(path)

    curtailment = parse_curtailment_daily(path, sheet="Detail2")

    assert curtailment.to_dict("records") == [
        {
            "source_sheet": "Detail2",
            "date": datetime(2025, 1, 31).date(),
            "curtailed_energy_kwh": 47406.89,
        }
    ]


def test_parse_dcm_outage_limitation_extracts_intervals(tmp_path: Path) -> None:
    path = tmp_path / "dcm.xlsx"
    _dcm_workbook(path)

    intervals = parse_dcm_outage_limitation(path, sheet="Juli 2025")

    outage = intervals[intervals["kind"] == "outage"].iloc[0]
    assert outage["date"] == datetime(2025, 7, 1).date()
    assert outage["start_time"] == "06:00:00"
    assert outage["restoration_time"] == "15:04:00"
    assert outage["equipment"] == "OG.3"
    assert outage["minute_lost"] == 544
    limitation = intervals[intervals["kind"] == "limitation"].iloc[0]
    assert limitation["equipment"] == "STS 6 - STS 10"
    assert limitation["restoration_time"] == "17:03:00"
    # the empty day produced no interval rows
    assert len(intervals) == 2


def test_file_provenance_reports_size_and_hash(tmp_path: Path) -> None:
    path = tmp_path / "maintenance.xlsx"
    _maintenance_workbook(path)

    provenance = file_provenance(path)

    assert provenance["source_name"] == "maintenance.xlsx"
    assert provenance["byte_size"] == path.stat().st_size
    assert len(provenance["sha256"]) == 64
